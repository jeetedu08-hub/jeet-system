import os
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
from matplotlib.backends.backend_pdf import PdfPages
import textwrap
import matplotlib.font_manager as fm
import matplotlib.patheffects as path_effects
import matplotlib.patches as mpatches
import traceback
import streamlit as st
import io
from supabase import create_client, Client
import zipfile
import re
import time 

# 엑셀 스타일링을 위한 라이브러리
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# --- 1. 환경 및 폰트 설정 ---
font_path = "malgun.ttf"
try:
    if os.path.exists(font_path):
        fm.fontManager.addfont(font_path)
        font_prop = fm.FontProperties(fname=font_path)
        plt.rcParams['font.family'] = font_prop.get_name()
    else:
        plt.rcParams['font.family'] = 'Malgun Gothic'
except Exception as e:
    print(f"폰트 로딩 실패 (기본 폰트로 대체됨): {e}")
    plt.rcParams['font.family'] = 'sans-serif'

plt.rcParams['axes.unicode_minus'] = False

COLOR_NAVY = '#1A237E'
COLOR_RED = '#D32F2F'
# ── 수정된 색상 팔레트 ──────────────────────────────────────
COLOR_STUDENT = '#1565C0'   # 학생: 파랑 (막대그래프)
COLOR_UNIT    = '#E65100'   # 반 평균: 진한 주황
COLOR_AVG     = '#2E7D32'   # 과정 평균: 초록
# 방사형 전용 fill 색상 (투명도 포함이라 별도 관리)
COLOR_RADAR_STUDENT = '#1565C0'
COLOR_RADAR_UNIT    = '#E65100'
COLOR_RADAR_AVG     = '#2E7D32'
# ───────────────────────────────────────────────────────────
COLOR_GRID = '#E0E0E0'
COLOR_BG = '#F8F9FA'

# --- 🛠️ 2. Supabase 연동 및 캠퍼스 동적 스위칭 설정 ---
def init_supabase_dynamic():
    supabase_secrets = st.secrets.get("supabase", {})
    secret_campus_name = supabase_secrets.get("campus_name", "")

    if "yeongtong_supabase" in st.secrets or secret_campus_name == "영통캠퍼스":
        if "yeongtong_supabase" in st.secrets:
            url = st.secrets["yeongtong_supabase"]["url"]
            key = st.secrets["yeongtong_supabase"]["key"]
        else:
            url = supabase_secrets.get("url")
            key = supabase_secrets.get("key")
            
        campus_name = "영통캠퍼스"
        campus_config = {
            "title_text": "📊 JEET 영통캠퍼스 성적 통합 관리 시스템",
            "logo_file": "logo_yeongtong.png",
            "footer_campuses": [
                ("영통 캠퍼스 주소 및 문의처", "전화번호 및 상세 주소를 입력하세요")
            ]
        }
    else:
        if "jukjeon_supabase" in st.secrets:
            url = st.secrets["jukjeon_supabase"]["url"]
            key = st.secrets["jukjeon_supabase"]["key"]
        else:
            url = st.secrets.get("SUPABASE_URL", supabase_secrets.get("url"))
            key = st.secrets.get("SUPABASE_KEY", supabase_secrets.get("key"))
            
        campus_name = "죽전캠퍼스"
        campus_config = {
            "title_text": "📊 JEET 죽전캠퍼스 성적 통합 관리 시스템",
            "logo_file": "logo.png",
            "footer_campuses": [
                ("수지 캠퍼스: 276-8003", "풍덕천로 129번길 16-1"), 
                ("죽전 캠퍼스: 263-8003", "기흥구 죽현로 29"), 
                ("광교 캠퍼스: 257-8003", "영통구 혜명로 10")
            ]
        }
    return create_client(url, key), campus_name, campus_config

try:
    supabase, CURRENT_CAMPUS, CAMPUS_CFG = init_supabase_dynamic()
except Exception as e:
    st.error(f"데이터베이스 연결 실패: {e}")
    st.stop()


@st.cache_data(ttl=1)
def fetch_all_dataframes():
    info_res = supabase.table("test_info").select("*").limit(10000).execute()
    df_info = pd.DataFrame(info_res.data)
    
    results_res = supabase.table("student_results").select("*").limit(10000).execute()
    df_results = pd.DataFrame(results_res.data)
    
    if not df_results.empty:
        if 'id' in df_results.columns:
            df_results = df_results.sort_values(by='id', ascending=False)
        elif 'created_at' in df_results.columns:
            df_results = df_results.sort_values(by='created_at', ascending=False)
            
        df_results = df_results.fillna(0)
        
        def normalize_col(col_name):
            col_str = str(col_name).strip().split('.')[0]
            nums = re.findall(r'\d+', col_str)
            return nums[0] if nums else col_str
            
        meta_cols = ['id', 'created_at', '시험명', '이름', '반', '학교', '학년', '분기', '구분', '총점', '맞은개수_2점', '맞은개수_3점', '맞은개수_4점']
        new_columns = []
        for col in df_results.columns:
            if col in meta_cols:
                new_columns.append(col)
            else:
                new_columns.append(normalize_col(col))
        df_results.columns = new_columns

        if not df_info.empty:
            df_info['문항번호_정제'] = df_info['문항번호'].apply(normalize_col)
            weight_dict = {}
            for _, info_row in df_info.iterrows():
                t_name = str(info_row.get('시험명', '')).strip()
                q_num = str(info_row.get('문항번호_정제', '')).strip()
                try:
                    w_val = int(float(info_row.get('배점', 3)))
                except:
                    w_val = 3
                weight_dict[(t_name, q_num)] = w_val

            for idx, res_row in df_results.iterrows():
                current_test = str(res_row.get('시험명', '')).strip()
                test_weights = {k[1]: v for k, v in weight_dict.items() if k[0] == current_test}
                
                if test_weights:
                    c_2, c_3, c_4 = 0, 0, 0
                    calculated_total = 0  
                    
                    for q_col, weight in test_weights.items():
                        if q_col in df_results.columns:
                            val = res_row[q_col]
                            if str(val).strip().upper() in ['1', '1.0', 'O', '정답', 'TRUE']:
                                calculated_total += weight  
                                if weight == 2: c_2 += 1
                                elif weight == 3: c_3 += 1
                                elif weight == 4: c_4 += 1
                    
                    if pd.isna(res_row.get('맞은개수_2점')) or res_row.get('맞은개수_2점') == 0:
                        df_results.at[idx, '맞은개수_2점'] = c_2
                    if pd.isna(res_row.get('맞은개수_3점')) or res_row.get('맞은개수_3점') == 0:
                        df_results.at[idx, '맞은개수_3점'] = c_3
                    if pd.isna(res_row.get('맞은개수_4점')) or res_row.get('맞은개수_4점') == 0:
                        df_results.at[idx, '맞은개수_4점'] = c_4
                        
                    if pd.isna(res_row.get('총점')) or res_row.get('총점') == 0:
                        df_results.at[idx, '총점'] = calculated_total
        
    return df_info, df_results


# ── 공통 스케일 변환: 실제 0~100% → 표시 20~100% ──────────
def scale_to_display(raw_val):
    """실제 점수 비율(0~100)을 표시 범위(20~100)로 변환"""
    return 20 + float(raw_val) * 0.8

def scale_list(raw_list):
    return [scale_to_display(v) for v in raw_list]
# ────────────────────────────────────────────────────────────


# ── 반명 정규화 헬퍼: 대소문자/공백 무시 ──────────────────
def normalize_class_name(val):
    return str(val).strip().replace(" ", "").upper()
# ────────────────────────────────────────────────────────────


# --- 3. 공통 그래프 그리기 함수 ---
def draw_report_figure(fig, s_row, student_name, student_grade, selected_test,
                       cat_ratio, avg_cat_ratio, unit_data, unit_avg_data, unit_order,
                       class_cat_ratio=None, class_unit_avg_data=None):

    border = plt.Rectangle((0.015, 0.015), 0.97, 0.97, fill=False,
                            edgecolor=COLOR_RED, linewidth=5.0,
                            transform=fig.transFigure, zorder=10)
    fig.patches.append(border)

    if os.path.exists("logo.png"):
        logo_img = plt.imread("logo.png")
        logo_img_axes = fig.add_axes([0.80, 0.915, 0.15, 0.045], zorder=15)
        logo_img_axes.imshow(logo_img)
        logo_img_axes.axis('off')

    txt_jeet  = fig.text(0.31, 0.88, 'JEET', fontsize=42, fontweight='bold', color=COLOR_RED, ha='right')
    txt_title = fig.text(0.33, 0.88, '수학 능력 분석 리포트', fontsize=32, fontweight='bold', color=COLOR_NAVY, ha='left')
    
    student_class   = str(s_row.get('반', '')).strip()
    student_quarter = str(s_row.get('분기', '')).strip()
    class_text   = f"{student_class} | " if student_class   and student_class   not in ('0','0.0') else ""
    quarter_text = f" [{student_quarter}]" if student_quarter and student_quarter not in ('0','0.0') else ""
    
    info_text = (f"학교: {s_row.get('학교','')} | 학년: {student_grade} | "
                 f"{class_text}이름: {student_name} | 과정: {selected_test}{quarter_text}")
    txt_info = fig.text(0.5, 0.84, info_text, ha='center', fontsize=12, fontweight='bold', color='#222')

    for t_obj, fg in [(txt_jeet, COLOR_RED), (txt_title, COLOR_NAVY), (txt_info, '#222')]:
        t_obj.set_path_effects([path_effects.withStroke(linewidth=1.5, foreground=fg)])

    # ══════════════════════════════════════════════════════════
    # ▶ 방사형 그래프 — 채우기 효과 + 실선 (굵은 선 제거)
    # ══════════════════════════════════════════════════════════
    ax1 = fig.add_axes([0.10, 0.52, 0.32, 0.22], polar=True)

    all_cats      = cat_ratio.index.tolist()
    ordered_labels = (['계산력'] + [c for c in all_cats if c != '계산력']
                      if '계산력' in all_cats else all_cats)
    s_ordered = cat_ratio.reindex(ordered_labels)
    labels    = s_ordered.index.tolist()
    angles    = np.linspace(0, 2*np.pi, len(labels), endpoint=False).tolist() + [0]

    ax1.set_theta_direction(-1)
    ax1.set_theta_offset(np.pi / 2.0)

    # ── 학생: 파랑 실선(-) + 진한 채우기 ─────────────────────
    s_raw    = s_ordered.values.tolist()
    s_scaled = scale_list(s_raw) + [scale_to_display(s_raw[0])]
    ax1.plot(angles, s_scaled,
             color=COLOR_RADAR_STUDENT, linewidth=2.2, linestyle='-', zorder=4,
             label='학생 점수')
    ax1.fill(angles, s_scaled, color=COLOR_RADAR_STUDENT, alpha=0.28, zorder=3)

    # ── 시험지 전체 평균: 초록 파선(--) + 채우기 ──────────────
    avg_raw    = avg_cat_ratio.reindex(ordered_labels).fillna(0).values.tolist()
    avg_scaled = scale_list(avg_raw) + [scale_to_display(avg_raw[0])]
    ax1.plot(angles, avg_scaled,
             color=COLOR_RADAR_AVG, linewidth=1.8, linestyle='--', dashes=(5, 3), zorder=4,
             label='과정 평균')
    ax1.fill(angles, avg_scaled, color=COLOR_RADAR_AVG, alpha=0.13, zorder=2)

    # ── 같은 반 평균: 주황 일점쇄선(-.) + 채우기 ─────────────
    if class_cat_ratio is not None:
        cls_raw    = class_cat_ratio.reindex(ordered_labels).fillna(0).values.tolist()
        cls_scaled = scale_list(cls_raw) + [scale_to_display(cls_raw[0])]
        ax1.plot(angles, cls_scaled,
                 color=COLOR_RADAR_UNIT, linewidth=1.8, linestyle='-.', zorder=4,
                 label='반 평균')
        ax1.fill(angles, cls_scaled, color=COLOR_RADAR_UNIT, alpha=0.13, zorder=1)

    ax1_limit = max(50, min(115, max(s_scaled) + 10))
    ax1.set_ylim(0, ax1_limit)
    ax1.set_xticks(angles[:-1])
    ax1.set_xticklabels([])
    ax1.set_yticklabels([])

    for i, label_text in enumerate(labels):
        angle    = angles[i]
        dist_tb  = ax1_limit * 1.05
        dist_lr  = ax1_limit * 1.02
        ha, va, dist = (('center','bottom', dist_tb) if angle == 0 else
                        ('left',  'center', dist_lr) if 0 < angle < np.pi else
                        ('center','top',    dist_tb) if angle == np.pi else
                        ('right', 'center', dist_lr))
        if '문제\n해결력' in label_text:
            dist += ax1_limit * 0.08
            ha = 'left' if 0 < angle < np.pi else 'right'
        ax1.text(angle, dist, label_text,
                 fontsize=10, fontweight='bold', va=va, ha=ha, color=COLOR_NAVY)

    ax1.legend(loc='upper right', bbox_to_anchor=(1.35, 1.15),
               fontsize=7.5, framealpha=0.85, handlelength=2.5, handleheight=1.2)
    title1 = ax1.set_title("▶ 영역별 핵심 역량 지표 (%)",
                            pad=30, fontsize=14, fontweight='bold', color=COLOR_NAVY)
    title1.set_path_effects([path_effects.withStroke(linewidth=1, foreground=COLOR_NAVY)])

    # ══════════════════════════════════════════════════════════
    # ▶ 단원별 성취도 막대그래프 — 파랑 막대 + 20% 기본값 스케일
    # ══════════════════════════════════════════════════════════
    ax2 = fig.add_axes([0.55, 0.54, 0.35, 0.18])

    final_unit_data = unit_data.reindex(unit_order).fillna(0)
    x_pos     = np.arange(len(final_unit_data))
    bar_width = 0.45

    # 실제 점수 비율(0~100) 계산 후 20~100 범위로 스케일
    s_pct_raw = (final_unit_data['득점'] / final_unit_data['배점'] * 100).fillna(0)
    s_pct     = s_pct_raw.apply(scale_to_display)          # 표시용 스케일
    max_b_val = s_pct.max() if not s_pct.empty else 0
    ax2_limit = max(50, min(115, max_b_val + 15))

    # ── 학생 막대: 파랑 ────────────────────────────────────────
    ax2.bar(x_pos, s_pct, color=COLOR_STUDENT, alpha=0.85,
            width=bar_width, zorder=3, label='학생 점수')

    # ── 시험지 전체 평균: 초록 실선 + 원형 마커 ───────────────
    if unit_avg_data is not None and not unit_avg_data.empty:
        avg_unit_pct = []
        for u in final_unit_data.index:
            denom  = final_unit_data.loc[u, '배점'] if u in final_unit_data.index else 0
            numer  = unit_avg_data.loc[u, '평균득점'] if u in unit_avg_data.index else 0
            raw_pct = (numer / denom * 100) if denom > 0 else 0
            avg_unit_pct.append(scale_to_display(raw_pct))
        ax2.plot(x_pos, avg_unit_pct,
                 color=COLOR_AVG, linewidth=2.2, linestyle='-',
                 marker='o', markersize=6,
                 markerfacecolor='white', markeredgewidth=2,
                 zorder=5, label='과정 평균')

    # ── 같은 반 평균: 주황 실선 + 다이아몬드 마커 ─────────────
    if class_unit_avg_data is not None and not class_unit_avg_data.empty:
        cls_unit_pct = []
        for u in final_unit_data.index:
            denom  = final_unit_data.loc[u, '배점'] if u in final_unit_data.index else 0
            numer  = class_unit_avg_data.loc[u, '평균득점'] if u in class_unit_avg_data.index else 0
            raw_pct = (numer / denom * 100) if denom > 0 else 0
            cls_unit_pct.append(scale_to_display(raw_pct))
        ax2.plot(x_pos, cls_unit_pct,
                 color=COLOR_UNIT, linewidth=2.2, linestyle='-',
                 marker='D', markersize=5,
                 markerfacecolor='white', markeredgewidth=2,
                 zorder=5, label='반 평균')

    ax2.legend(loc='upper right', fontsize=7.5, framealpha=0.85, handlelength=2.5)
    ax2.set_xticks(x_pos)
    ax2.set_xticklabels([textwrap.fill(str(l), 5) for l in final_unit_data.index],
                        fontsize=8, fontweight='bold')
    ax2.tick_params(axis='x', which='both', length=0)

    # y축: 20(=0%) ~ ax2_limit, 눈금을 실제 점수 기준으로 표시
    ax2.set_ylim(20, ax2_limit)
    ytick_real  = [0, 20, 40, 60, 80, 100]
    ytick_disp  = [scale_to_display(v) for v in ytick_real]
    ytick_disp  = [v for v in ytick_disp if 20 <= v <= ax2_limit]
    ytick_label = [f"{int(round((v - 20) / 0.8))}%" for v in ytick_disp]
    ax2.set_yticks(ytick_disp)
    ax2.set_yticklabels(ytick_label, fontsize=7)
    ax2.grid(axis='y', color=COLOR_GRID, linestyle='-', linewidth=0.5, zorder=0)

    title2 = ax2.set_title("▶ 단원별 성취도 (%)",
                            pad=25, fontsize=14, fontweight='bold', color=COLOR_NAVY)
    title2.set_path_effects([path_effects.withStroke(linewidth=1, foreground=COLOR_NAVY)])

    # ══════════════════════════════════════════════════════════
    # ▶ 하단 심층 분석 박스
    # ══════════════════════════════════════════════════════════
    BOX_LEFT   = 0.05
    BOX_RIGHT  = 0.95
    BOX_TOP    = 0.470
    BOX_BOTTOM = 0.110

    # 외곽 박스
    rect_diag = plt.Rectangle(
        (BOX_LEFT, BOX_BOTTOM),
        BOX_RIGHT - BOX_LEFT, BOX_TOP - BOX_BOTTOM,
        fill=True, facecolor='#F5F6FA',
        edgecolor='#B0BEC5', linewidth=1.2,
        transform=fig.transFigure, zorder=0
    )
    fig.patches.append(rect_diag)

    # 헤더 타이틀 바
    HDR_H = 0.032
    header_rect = plt.Rectangle(
        (BOX_LEFT, BOX_TOP - HDR_H),
        BOX_RIGHT - BOX_LEFT, HDR_H,
        fill=True, facecolor=COLOR_NAVY,
        transform=fig.transFigure, zorder=1
    )
    fig.patches.append(header_rect)
    fig.text(0.50, BOX_TOP - HDR_H / 2,
             f"JEET 중등 자사 센터  |  {student_name} 학생 심층 분석",
             ha='center', va='center',
             fontsize=10.5, fontweight='bold', color='white', zorder=2)

    # ── 텍스트 내용 생성 ──────────────────────────────────────
    u_res   = s_pct_raw
    avg_val = int(cat_ratio.mean())

    if avg_val >= 80:
        eval_tier = "심화 개념의 완벽한 체득과 날카로운 수학적 직관력을 겸비한 최상위권 수준의 성취를 보여줍니다. 고난도 문항 해결 능력이 탁월하며 자기 주도적 심화 학습이 충분히 가능한 상태"
    elif avg_val >= 60:
        eval_tier = "탄탄한 기본기를 바탕으로 성실하게 학습의 밀도를 높여가는 우수한 성취를 보여줍니다. 안정적인 개념 활용 능력을 갖추고 있어, 향후 고난도 유형에 대한 도전이 성장의 핵심이 될 수 있는 시점"
    elif avg_val >= 20:
        eval_tier = "핵심 개념을 정교하게 다듬어가는 과정에 있으며, 학습 잠재력이 점진적으로 발현되는 도약 단계의 성취를 보여줍니다. 꾸준한 학습 태도를 유지한다면 성취도 향상이 기대되는 상황"
    else:
        eval_tier = "수학적 기초 체력을 보강하며 자신감을 쌓아가는 기틀 마련 단계의 성취를 보여줍니다. 학습적 결손을 메우고 성공적인 문제 풀이 경험을 축적하여 학습 동기를 부여하는 데 집중해야 하는 시점"

    diag_total = f"{student_name} 학생은 성취도 {avg_val}%를 기록하며, 현재 {eval_tier}입니다."

    c_best = cat_ratio[cat_ratio >= 80].index.str.replace('\n','').tolist()
    c_good = cat_ratio[(cat_ratio >= 50) & (cat_ratio < 80)].index.str.replace('\n','').tolist()
    c_weak = cat_ratio[(cat_ratio >= 20) & (cat_ratio < 50)].index.str.replace('\n','').tolist()
    c_warn = cat_ratio[cat_ratio < 20].index.tolist()

    diag_combined = ""
    if c_best: diag_combined += f"특히 {', '.join([f'[{c}]' for c in c_best])} 영역에서 정교한 논리 구조와 높은 응용력을 보이며 압도적인 강점을 나타냅니다. 현재의 감각을 유지하며 최고 난도 문항을 통해 사고의 확장을 이어가야 합니다. "
    if c_good: diag_combined += f"{', '.join([f'[{c}]' for c in c_good])} 영역은 안정적인 정답률로 견고한 기본기를 증명하였습니다. 다만, 문항의 복합도가 높아질 때 발생하는 오답을 줄이기 위해 심화 유형에 대한 반복 훈련이 병행되어야 합니다. "
    if c_weak: diag_combined += f"{', '.join([f'[{c}]' for c in c_weak])} 영역은 복합 개념을 문제에 투영하는 과정에서 다소 고전하는 모습이 보입니다. 단편적인 문제 풀이보다는 개념 간의 유기적 연결성을 이해하고 유사 유형을 집중 분석하는 보완 작업이 필요합니다. "
    if c_warn: diag_combined += f"{', '.join([f'[{c}]' for c in c_warn])} 영역은 단원 간 연계성이 높은 만큼, 기초 개념의 재정립과 필수 유형에 대한 집중 학습이 필요합니다. 단계별 맞춤 문항을 통해 이해도를 근본적으로 끌어올려 실력을 끌어올릴 필요가 있습니다. "

    g_best = u_res[u_res >= 80].index.tolist()
    g_good = u_res[(u_res >= 50) & (u_res < 80)].index.tolist()
    g_weak = u_res[(u_res >= 20) & (u_res < 50)].index.tolist()
    g_warn = u_res[u_res < 20].index.tolist()

    if g_best: diag_combined += f"{', '.join([f'<{u}>' for u in g_best])} 단원은 기본 및 응용 단계를 넘어 심화 단계까지 완벽히 소화하고 있습니다. 이제는 일반적인 유형 학습보다는 사고의 폭을 넓힐 수 있는 '킬러 문항' 중심의 도전적 학습이 필요한 단계입니다. "
    if g_good: diag_combined += f"{', '.join([f'<{u}>' for u in g_good])} 단원은 필수 유형들은 막힘없이 해결하고 있습니다. 이제는 단편적인 문제 풀이에서 벗어나, 개념을 다각도로 비트는 응용 문항에 대한 적응력을 키워야 할 때입니다. 심화 문항 도전 횟수를 늘려 사고의 유연성을 기른다면 보다 높은 수학 실력을 갖출 수 있을 것으로 판단됩니다. "
    if g_weak: diag_combined += f"{', '.join([f'<{u}>' for u in g_weak])} 단원은 개념의 조각들은 인지하고 있으나 실전 적용에서 병목 현상이 관찰됩니다. 난이도 높은 문제를 해결하기보다 핵심 유형의 반복 숙달이 필요하고, 성공적인 문제 풀이 경험을 축적하여 해당 단원의 수학적 자신감을 높인다면 한단계 더 발전할 수 있을 것으로 판단됩니다. "
    if g_warn: diag_combined += f"{', '.join([f'<{u}>' for u in g_warn])} 단원은 현재는 잠재력이 발현되기 전의 응축 단계입니다. 수학적 기초 체력이 부족할 뿐, 적절한 자극과 맞춤형 관리가 병행된다면 충분히 반등할 수 있는 가능성을 가지고 있습니다. 아이가 포기하지 않도록 가정에서 따뜻한 격려 부탁드립니다. "
    if not (g_best or g_good or g_weak or g_warn):
        diag_combined += "전반적인 단원별 성취도가 매우 균형 있게 나타나고 있습니다. 어느 한 쪽으로 치우치지 않는 고른 학습 균형이 큰 강점입니다."

    weak_list  = u_res[u_res < 40].index.tolist()
    avg_score  = u_res.mean()
    units_text = ', '.join([f'<{u}>' for u in weak_list]) if weak_list else "핵심"

    if avg_score >= 80:
        sol_text = f"{student_name} 학생은 훌륭한 성취를 보여주고 있습니다. JEET CARE+를 통해 고난도 사고력 문제를 즐길 수 있는 환경을 만들고, JDM으로 자기주도학습 습관을 형성하며 주말몰입수업으로 심화를 다지겠습니다."
    elif avg_score >= 50:
        sol_text = f"배운 내용을 자기 것으로 만드는 과정이 순조롭습니다. JEET CARE+를 통해 응용·심화 문제를 연습하고, JDM으로 자기주도학습 습관을 기르며 주말몰입수업으로 성취도를 끌어올리겠습니다."
    elif avg_score >= 20:
        sol_text = f"지금은 한 단계 더 성장하기 위해 에너지를 모으는 시기입니다. {units_text} 단원을 JEET CARE+ 밀착 관리로 친숙한 개념으로 바꿔가겠습니다. JDM으로 자기주도학습 습관을 만들고 주말몰입수업으로 수학이 즐거운 과목이 되도록 지도하겠습니다."
    else:
        sol_text = f"기초를 단단히 다지는 것이 가장 확실한 성장의 길입니다. {units_text} 단원의 기초를 JEET CARE+와 JDM으로 다지며, 주말몰입수업과 눈높이 지도로 작은 노력이 큰 성취로 이어지도록 정성을 다하겠습니다."

    # ── 레이아웃 상수 ─────────────────────────────────────────
    # figure 좌표: 0.0 ~ 1.0 (A4 세로 기준 실측)
    # 가용 콘텐츠 영역
    CL = BOX_LEFT  + 0.022          # Content Left  (패딩)
    CR = BOX_RIGHT - 0.022          # Content Right (패딩)
    CT = BOX_TOP   - HDR_H - 0.010  # Content Top   (헤더 바 아래 + 여백)
    CB = BOX_BOTTOM + 0.010         # Content Bottom

    # figure 좌표에서 텍스트 폰트 크기 환산:
    # A4 세로(11.69인치) × dpi(300) ÷ 72 pt/inch = 1pt ≈ 0.000285 figure 단위
    # 한글 폰트는 영문보다 약 1.5배 넓으므로 wrap 너비를 보수적으로 잡음
    PT_TO_FIG = 1.0 / (11.69 * 72)   # 1pt → figure 단위 높이

    section_defs = [
        ("[종합 진단]",          diag_total,    '#1A237E', '#E8EAF6'),
        ("[영역별&단원별 분석]", diag_combined, '#1B5E20', '#E8F5E9'),
        ("[JEET 맞춤 솔루션]",   sol_text,      '#B71C1C', '#FFEBEE'),
    ]

    # 전체 글자 수에 따라 폰트 크기 결정
    total_chars = sum(len(c) for _, c, _, _ in section_defs)
    if total_chars > 650:
        fs_body = 6.5
    elif total_chars > 450:
        fs_body = 7.0
    else:
        fs_body = 7.5
    fs_tag    = fs_body + 0.8

    # 줄 높이: 폰트 포인트 × 줄간격(1.55) → figure 단위
    LINE_H    = fs_body * PT_TO_FIG * 1.55
    TAG_H     = fs_tag  * PT_TO_FIG * 2.2   # 태그 배지 높이
    TAG_MGAP  = fs_body * PT_TO_FIG * 0.9   # 태그~본문 간격
    SEC_GAP   = fs_body * PT_TO_FIG * 2.0   # 섹션 간 간격

    # figure 너비(8.27인치) 기준으로 한 줄에 들어갈 한글 글자 수 계산
    # CL~CR 사이 인치 폭 / (글자 1개 인치 폭)
    fig_width_inch  = 8.27
    char_w_inch     = fs_body / 72 * 1.05   # 한글 약 1em 너비 (pt → inch × 여유)
    avail_w_inch    = (CR - CL) * fig_width_inch
    wrap_w          = max(30, int(avail_w_inch / char_w_inch))

    curr_y = CT

    for idx, (tag, content, tag_color, tag_bg) in enumerate(section_defs):
        if curr_y < CB + TAG_H + LINE_H:
            break

        # ── 섹션 구분선 (첫 섹션 제외) ───────────────────────
        if idx > 0 and curr_y < CT - 0.01:
            sep = plt.Line2D(
                [CL, CR], [curr_y + SEC_GAP * 0.3, curr_y + SEC_GAP * 0.3],
                color='#C5CAE9', linewidth=0.7,
                transform=fig.transFigure, zorder=2
            )
            fig.lines.append(sep)

        # ── 태그 배지: bbox 속성으로 텍스트 크기에 딱 맞게 자동 생성 ─
        fig.text(
            CL, curr_y - TAG_H * 0.35,
            tag,
            fontsize=fs_tag, fontweight='bold', color=tag_color,
            va='center', ha='left', zorder=3,
            bbox=dict(
                boxstyle='round,pad=0.25',
                facecolor=tag_bg,
                edgecolor=tag_color,
                linewidth=0.9,
            )
        )

        curr_y -= TAG_H + TAG_MGAP

        # ── 본문 텍스트 ───────────────────────────────────────
        wrapped_lines = textwrap.fill(content, width=wrap_w).split('\n')
        for line in wrapped_lines:
            if curr_y < CB:
                break
            fig.text(CL, curr_y, line,
                     fontsize=fs_body, color='#2C2C2C',
                     va='top', zorder=3)
            curr_y -= LINE_H

        curr_y -= SEC_GAP

    # ── 하단 캠퍼스 안내 ──────────────────────────────────────
    line_footer = plt.Line2D([0.05, 0.95], [0.10, 0.10],
                             color=COLOR_NAVY, linewidth=1,
                             transform=fig.transFigure)
    fig.lines.append(line_footer)

    campuses = CAMPUS_CFG["footer_campuses"]
    pos_x = [0.50] if len(campuses) == 1 else [0.22, 0.50, 0.78]
    for i, (name, addr) in enumerate(campuses):
        fig.text(pos_x[i], 0.07, name,  ha='center', fontsize=10, fontweight='bold', color=COLOR_NAVY)
        fig.text(pos_x[i], 0.045, addr, ha='center', fontsize=7.5, color='#555')


# --- 4. 데이터 가공 헬퍼 함수 ---
def prepare_report_data(selected_test):
    df_info, df_results = fetch_all_dataframes()
    
    selected_test_clean = str(selected_test).strip()
    
    df_info['시험명_정제'] = df_info['시험명'].astype(str).str.strip()
    df_results['시험명_정제'] = df_results['시험명'].astype(str).str.strip()
    
    df_info    = df_info[df_info['시험명_정제'] == selected_test_clean].copy()
    df_results = df_results[df_results['시험명_정제'] == selected_test_clean].copy()
    
    df_info['배점']  = pd.to_numeric(df_info['배점'], errors='coerce').fillna(3).astype(int)
    df_info['단원']  = df_info['단원'].astype(str).str.strip()
    df_info['영역']  = df_info['영역'].astype(str).str.strip()
    df_info['영역']  = df_info['영역'].str.replace('문제해결력', '문제\n해결력')
    
    def clean_info_q(q):
        nums = re.findall(r'\d+', str(q).split('.')[0])
        return nums[0] if nums else str(q).strip()
    df_info['문항번호'] = df_info['문항번호'].apply(clean_info_q)
    
    unit_order = df_info['단원'].drop_duplicates().tolist()
    q_cols     = df_info['문항번호'].tolist()
    
    def safe_to_binary(val):
        if hasattr(val, 'any') or hasattr(val, 'all'): return 0
        if pd.isna(val): return 0
        v_str = str(val).strip().upper()
        if v_str in ['O','1','1.0','정답','TRUE']: return 1
        if v_str in ['X','0','0.0','오답','FALSE','']: return 0
        try: return 1 if float(val) > 0 else 0
        except: return 0

    if not df_results.empty:
        df_scores = pd.DataFrame(index=df_results.index, columns=q_cols)
        for q in q_cols:
            if q in df_results.columns:
                target_col = df_results[q]
                if isinstance(target_col, pd.DataFrame):
                    target_col = target_col.iloc[:, 0]
                df_scores[q] = target_col.apply(safe_to_binary)
            else:
                df_scores[q] = 0
    else:
        df_scores = pd.DataFrame(0, index=[0], columns=q_cols)

    avg_per_q      = df_scores.mean()
    total_analysis = df_info.copy()
    total_analysis['평균득점'] = total_analysis['문항번호'].apply(lambda x: avg_per_q.get(str(x), 0))
    
    avg_cat_ratio = (total_analysis.groupby('영역')['평균득점'].sum() /
                     total_analysis.groupby('영역')['배점'].sum() * 100).fillna(0)
    unit_avg_data = total_analysis.groupby('단원').agg({'평균득점': 'sum'})
    unit_avg_data = unit_avg_data.reindex([u for u in unit_order if u in unit_avg_data.index])
    
    return df_info, df_results, avg_cat_ratio, unit_avg_data, unit_order, safe_to_binary, total_analysis


def generate_jeet_expert_report(target_name, selected_test):
    try:
        df_info, df_results, avg_cat_ratio, unit_avg_data, unit_order, safe_to_binary, total_analysis = prepare_report_data(selected_test)
        student_found = False
        img_buffer    = io.BytesIO()
        
        for _, s_row in df_results.iterrows():
            student_name = str(s_row.get('이름', '')).strip()
            db_name      = student_name.replace(" ", "").upper()
            search_name  = str(target_name).replace(" ", "").strip().upper()
            if not db_name or db_name == '0' or db_name != search_name:
                continue
                
            student_found = True
            student_grade = s_row.get('학년', '')
            
            analysis = df_info.copy()
            student_answers = []
            for q in analysis['문항번호']:
                val = s_row.get(str(q), None)
                if isinstance(val, pd.Series): val = val.iloc[0]
                student_answers.append(safe_to_binary(val) if val is not None else 0)
                
            analysis['정답여부'] = student_answers
            analysis['득점']     = analysis['정답여부'] * analysis['배점']
            
            cat_ratio = (analysis.groupby('영역')['득점'].sum() /
                         analysis.groupby('영역')['배점'].sum() * 100).fillna(0)
            unit_data = analysis.groupby('단원').agg({'득점':'sum', '배점':'sum'})
            unit_data = unit_data.reindex(unit_order).fillna(0)

            # ── 반 평균: 학년 + 반명(대소문자/공백 무시) 모두 일치하는 학생끼리만 ──
            student_class_norm = normalize_class_name(s_row.get('반', ''))
            student_grade_s    = str(s_row.get('학년', '')).strip()
            class_students = df_results[
                (df_results['반'].astype(str).apply(normalize_class_name) == student_class_norm) &
                (df_results['학년'].astype(str).str.strip() == student_grade_s)
            ]
            class_cat_ratio = class_unit_avg_data = None
            if len(class_students) > 1:
                cls_analysis = df_info.copy()
                cls_scores   = pd.DataFrame(index=class_students.index,
                                            columns=cls_analysis['문항번호'].tolist())
                for q in cls_analysis['문항번호'].tolist():
                    cls_scores[q] = class_students[q].apply(safe_to_binary) if q in class_students.columns else 0
                cls_avg_per_q = cls_scores.mean()
                cls_total     = cls_analysis.copy()
                cls_total['평균득점'] = cls_total['문항번호'].apply(lambda x: cls_avg_per_q.get(str(x), 0))
                class_cat_ratio = (cls_total.groupby('영역')['평균득점'].sum() /
                                   cls_total.groupby('영역')['배점'].sum() * 100).fillna(0)
                class_unit_avg_data = cls_total.groupby('단원').agg({'평균득점':'sum'})
                class_unit_avg_data = class_unit_avg_data.reindex(
                    [u for u in unit_order if u in class_unit_avg_data.index])

            fig = plt.figure(figsize=(8.27, 11.69))
            draw_report_figure(fig, s_row, student_name, student_grade, selected_test,
                               cat_ratio, avg_cat_ratio, unit_data, unit_avg_data, unit_order,
                               class_cat_ratio, class_unit_avg_data)
            fig.savefig(img_buffer, format='png', dpi=300, bbox_inches='tight')
            plt.close(fig)
            break
            
        if not student_found: return False, None, "학생을 찾을 수 없습니다."
        img_buffer.seek(0)
        return True, img_buffer, "리포트 생성 완료!"
    except Exception as e:
        return False, None, f"오류 발생: {traceback.format_exc()}"


def generate_batch_report(target_class, selected_test, selected_students=None):
    try:
        df_info, df_results, avg_cat_ratio, unit_avg_data, unit_order, safe_to_binary, total_analysis = prepare_report_data(selected_test)
        
        class_students = df_results[df_results['반'].astype(str).str.strip() == str(target_class).strip()]
        class_students = class_students.drop_duplicates(subset=['이름'], keep='first')
        
        if selected_students is not None:
            cleaned_selected = [str(s).strip() for s in selected_students]
            class_students = class_students[class_students['이름'].astype(str).str.strip().isin(cleaned_selected)]
        
        if class_students.empty:
            return False, None, "선택된 학생 데이터가 없습니다."
            
        zip_buffer = io.BytesIO()
        
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
            for _, s_row in class_students.iterrows():
                student_name = str(s_row.get('이름', '')).strip()
                if not student_name or student_name == '0': continue
                student_grade = s_row.get('학년', '')
                
                analysis = df_info.copy()
                student_answers = []
                for q in analysis['문항번호']:
                    val = s_row.get(str(q), None)
                    if isinstance(val, pd.Series): val = val.iloc[0]
                    student_answers.append(safe_to_binary(val) if val is not None else 0)
                
                analysis['정답여부'] = student_answers
                analysis['득점']     = analysis['정답여부'] * analysis['배점']
                cat_ratio = (analysis.groupby('영역')['득점'].sum() /
                             analysis.groupby('영역')['배점'].sum() * 100).fillna(0)
                unit_data = analysis.groupby('단원').agg({'득점':'sum','배점':'sum'})
                unit_data = unit_data.reindex(unit_order).fillna(0)

                # ── 반 평균: 학년 + 반명(대소문자/공백 무시) 모두 일치하는 학생끼리만 ──
                student_class_norm = normalize_class_name(s_row.get('반', ''))
                student_grade_s    = str(s_row.get('학년', '')).strip()
                same_class    = df_results[
                    (df_results['반'].astype(str).apply(normalize_class_name) == student_class_norm) &
                    (df_results['학년'].astype(str).str.strip() == student_grade_s)
                ]
                class_cat_ratio = class_unit_avg_data = None
                if len(same_class) > 1:
                    cls_analysis = df_info.copy()
                    cls_scores   = pd.DataFrame(index=same_class.index,
                                                columns=cls_analysis['문항번호'].tolist())
                    for q in cls_analysis['문항번호'].tolist():
                        cls_scores[q] = same_class[q].apply(safe_to_binary) if q in same_class.columns else 0
                    cls_avg_per_q = cls_scores.mean()
                    cls_total     = cls_analysis.copy()
                    cls_total['평균득점'] = cls_total['문항번호'].apply(lambda x: cls_avg_per_q.get(str(x), 0))
                    class_cat_ratio = (cls_total.groupby('영역')['평균득점'].sum() /
                                       cls_total.groupby('영역')['배점'].sum() * 100).fillna(0)
                    class_unit_avg_data = cls_total.groupby('단원').agg({'평균득점':'sum'})
                    class_unit_avg_data = class_unit_avg_data.reindex(
                        [u for u in unit_order if u in class_unit_avg_data.index])

                fig = plt.figure(figsize=(8.27, 11.69), dpi=300)
                draw_report_figure(fig, s_row, student_name, student_grade, selected_test,
                                   cat_ratio, avg_cat_ratio, unit_data, unit_avg_data, unit_order,
                                   class_cat_ratio, class_unit_avg_data)
                
                temp_buf = io.BytesIO()
                fig.savefig(temp_buf, format='png', dpi=300, bbox_inches='tight', orientation='portrait')
                plt.clf(); plt.close(fig)
                
                zip_file.writestr(f"{student_name}_리포트.png", temp_buf.getvalue())
            
        zip_buffer.seek(0)
        return True, zip_buffer, f"'{target_class}' 반 리포트 일괄 생성 완료!"
    except Exception as e:
        return False, None, f"오류 발생: {traceback.format_exc()}"


def export_excel_styled(df, quarter_name, q_cols):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "성적데이터"
    
    navy_fill  = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
    bg_color_1 = PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid")
    bg_color_2 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    white_font  = Font(name="Malgun Gothic", size=10, bold=True, color="FFFFFF")
    normal_font = Font(name="Malgun Gothic", size=10, bold=False, color="000000")
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'), right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin',  color='D3D3D3'), bottom=Side(style='thin', color='D3D3D3')
    )
    
    df['반_정제'] = df['반'].astype(str).str.strip()
    df_sorted    = df.sort_values(by=['반_정제','이름'])
    headers = ["시험명","구분","이름","반","학교","학년","분기","총점",
               "맞은개수_2점","맞은개수_3점","맞은개수_4점"]

    for c_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c_idx, value=header)
        cell.fill = navy_fill; cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    current_class = None; color_toggle = True
    for r_idx, (_, row) in enumerate(df_sorted.iterrows(), 2):
        if row['반_정제'] != current_class:
            current_class = row['반_정제']; color_toggle = not color_toggle
        current_fill = bg_color_1 if color_toggle else bg_color_2
        for c_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=row.get(header,""))
            cell.font = normal_font; cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = current_fill

    for col in ws.columns:
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=0)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    excel_buffer = io.BytesIO()
    wb.save(excel_buffer); excel_buffer.seek(0)
    return excel_buffer


# --- 5. Streamlit UI 레이아웃 구성 ---
st.set_page_config(page_title="JEET 통합 관리 시스템", layout="wide", page_icon="📊")

if st.sidebar.button("🔄 데이터베이스 새로고침", use_container_width=True):
    st.cache_data.clear()
    st.success("데이터를 새로 불러왔습니다!")
    st.rerun()

st.sidebar.markdown("---")
st.sidebar.markdown(f"### 🏢 현재 접속: **{CURRENT_CAMPUS}**")

col1, col2 = st.columns([8, 2])
with col1:
    st.title(CAMPUS_CFG["title_text"])
with col2:
    logo_file = CAMPUS_CFG["logo_file"]
    if os.path.exists(logo_file): st.image(logo_file, width=150)

try:
    df_info_all, df_results_all = fetch_all_dataframes()
except Exception as e:
    st.error(f"데이터베이스 로드 실패: {e}"); st.stop()

st.sidebar.header("📚 시험 과정 선택")
if not df_info_all.empty and '시험명' in df_info_all.columns:
    test_list = df_info_all['시험명'].dropna().unique().tolist()
else:
    test_list = []
    
if not test_list:
    st.warning("데이터베이스에 시험 정보가 없습니다."); st.stop()

selected_test       = st.sidebar.selectbox("분석할 시험 과정을 선택하세요:", test_list)
df_info_filtered    = df_info_all[df_info_all['시험명'] == selected_test]

tab1, tab2, tab3, tab4, tab5 = st.tabs([
    "✍️ 성적 입력",
    "📊 개별 리포트 출력",
    "📚 반별 일괄 리포트 출력",
    "🟢 분기별 엑셀 추출",
    "✏️ 성적 수정/삭제"
])

# --- Tab 1: 성적 입력 ---
with tab1:
    st.subheader(f"[{selected_test}] 학생 성적 입력")
    
    if not df_info_filtered.empty:
        q_weight_map = dict(zip(df_info_filtered['문항번호'].astype(str),
                                df_info_filtered['배점'].astype(int)))
        question_numbers = sorted(list(q_weight_map.keys()),
                                  key=lambda x: int(re.findall(r'\d+', x)[0]) if re.findall(r'\d+', x) else x)
    else:
        q_weight_map = {}; question_numbers = []

    if question_numbers:
        if "input_session_key" not in st.session_state:
            st.session_state["input_session_key"] = 0
        sk = st.session_state["input_session_key"]

        ci1,ci2,ci3,ci4,ci5,ci6 = st.columns([1.2,1.5,1.5,1.5,1.2,1.5])
        with ci1: input_type    = st.radio("구분", ["재원생","신규생"], key=f"input_type_{sk}", horizontal=True)
        with ci2: input_name    = st.text_input("이름", key=f"input_name_{sk}")
        with ci3: input_class   = st.text_input("반 (예: A반)", key=f"input_class_{sk}")
        with ci4: input_school  = st.text_input("학교", key=f"input_school_{sk}")
        with ci5: input_grade   = st.selectbox("학년", ["중1","중2","중3"], key=f"input_grade_{sk}")
        with ci6: input_quarter = st.selectbox(
            "분기 선택",
            ["2025년 4분기","2026년 1분기","2026년 2분기","2026년 3분기","2026년 4분기",
             "2027년 1분기","2027년 2분기","기타/정기 평가"],
            key=f"input_quarter_{sk}"
        )
        
        st.markdown("---")
        answers = {}
        for i in range(0, len(question_numbers), 4):
            cols = st.columns(4)
            for j, q_num in enumerate(question_numbers[i:i+4]):
                with cols[j]:
                    choice = st.radio(f"**{q_num}번 ({q_weight_map[q_num]}점)**",
                                      options=["O","X"], horizontal=True, key=f"q_{q_num}_{sk}")
                    answers[str(q_num)] = q_weight_map[q_num] if choice == "O" else 0

        st.markdown("---")
        total_score = sum(answers.values())
        count_2pt = sum(1 for q,s in answers.items() if q_weight_map[q]==2 and s>0)
        count_3pt = sum(1 for q,s in answers.items() if q_weight_map[q]==3 and s>0)
        count_4pt = sum(1 for q,s in answers.items() if q_weight_map[q]==4 and s>0)
        count_etc = sum(1 for q,s in answers.items() if q_weight_map[q] not in [2,3,4] and s>0)

        st.markdown("### 📈 실시간 채점 결과 요약")
        sc1,sc2,sc3,sc4,sc5 = st.columns(5)
        with sc1: st.metric("💯 현재 총점",     f"{total_score} 점")
        with sc2: st.metric("🟢 2점 맞은 개수", f"{count_2pt} 개")
        with sc3: st.metric("🔵 3점 맞은 개수", f"{count_3pt} 개")
        with sc4: st.metric("🔴 4점 맞은 개수", f"{count_4pt} 개")
        with sc5:
            if count_etc > 0: st.metric("🟡 기타 배점 정답", f"{count_etc} 개")
            else:              st.metric("📝 총 문항 수",     f"{len(question_numbers)} 문항")

        st.markdown("---")
        if st.button("DB에 성적 저장하기", type="primary", use_container_width=True):
            clean_name = input_name.strip()
            if not clean_name:
                st.error("⚠ 이름을 입력해주세요.")
            else:
                try:
                    new_record = {
                        "시험명": selected_test, "구분": input_type,
                        "이름": clean_name, "반": input_class,
                        "학교": input_school, "학년": input_grade,
                        "분기": input_quarter, "총점": total_score,
                        "맞은개수_2점": count_2pt, "맞은개수_3점": count_3pt,
                        "맞은개수_4점": count_4pt
                    }
                    for q_num in question_numbers:
                        new_record[str(q_num)] = 1 if answers[str(q_num)] > 0 else 0
                    supabase.table("student_results").insert(new_record).execute()
                    st.cache_data.clear()
                    st.success(f"🎉 [{input_type}] {clean_name} 학생의 [{input_quarter}] 성적({total_score}점)이 저장되었습니다!")
                    st.session_state["input_session_key"] += 1
                    time.sleep(2.0); st.rerun()
                except Exception as e:
                    st.error(f"저장 중 오류 발생: {e}")

# --- Tab 2: 개별 리포트 출력 ---
with tab2:
    st.subheader(f"[{selected_test}] 개별 심층 분석 리포트 생성")
    target_student = st.text_input("리포트를 출력할 학생 이름:", placeholder="예: 홍길동")
    if st.button("개별 리포트 생성 (PNG)", type="primary"):
        with st.spinner("리포트 생성 중..."):
            success, buf, msg = generate_jeet_expert_report(target_student.strip(), selected_test)
            if success:
                st.success(msg)
                st.download_button("🖼️ 이미지(PNG) 다운로드", buf.getvalue(),
                                   f"{target_student}_리포트.png", "image/png")
            else: st.error(msg)

# --- Tab 3: 반별 일괄 리포트 출력 ---
with tab3:
    st.subheader("📅 분기별/반별 전체 심층 분석 일괄 출력")
    
    if '분기' in df_results_all.columns and not df_results_all.empty:
        all_quarters = df_results_all['분기'].astype(str).str.strip().unique().tolist()
        quarter_list = sorted([q for q in all_quarters if q and q not in ['0','0.0','nan','None','']], reverse=True)
        
        if quarter_list:
            selected_quarter = st.selectbox("📅 출력할 분기를 선택하세요:", quarter_list, key="batch_quarter_select")
            df_quarter_filtered = df_results_all[df_results_all['분기'].astype(str).str.strip() == selected_quarter]
            
            all_tests_in_quarter = df_quarter_filtered['시험명'].astype(str).str.strip().unique().tolist()
            test_list_batch = sorted([t for t in all_tests_in_quarter if t and t not in ['0','nan','None','']])
            
            if test_list_batch:
                selected_test_batch = st.selectbox(
                    f"📝 [{selected_quarter}]의 시험 과정을 선택하세요:",
                    test_list_batch, key="batch_test_select_under_quarter"
                )
                df_final_filtered = df_quarter_filtered[
                    df_quarter_filtered['시험명'].astype(str).str.strip() == selected_test_batch
                ]
                
                all_grades = df_final_filtered['학년'].astype(str).str.strip().unique().tolist()
                grade_list = sorted([g for g in all_grades if g and g not in ['0','0.0','nan','None','']])
                if grade_list:
                    selected_grade_batch = st.selectbox("🎓 학년을 선택하세요:", grade_list, key="batch_grade_select")
                    df_grade_filtered = df_final_filtered[
                        df_final_filtered['학년'].astype(str).str.strip() == selected_grade_batch
                    ]
                else:
                    st.warning("⚠ 학년 데이터가 존재하지 않습니다.")
                    df_grade_filtered = df_final_filtered; selected_grade_batch = ""

                all_classes = df_grade_filtered['반'].astype(str).str.strip().unique().tolist()
                class_list  = sorted([c for c in all_classes if c and c not in ['0','0.0','nan','None','']])
                if class_list:
                    target_class = st.selectbox("📌 출력할 반을 선택하세요:", class_list, key="batch_class_select")
                    students_in_class = df_grade_filtered[
                        df_grade_filtered['반'].astype(str).str.strip() == target_class
                    ]['이름'].astype(str).str.strip().unique().tolist()
                    students_in_class = sorted([s for s in students_in_class if s and s not in ['0','nan','None','']])
                    if students_in_class:
                        selected_students = st.multiselect(
                            "👇 출력할 학생을 선택하세요:",
                            options=students_in_class, default=students_in_class,
                            key="batch_student_select"
                        )
                    else:
                        st.warning(f"⚠ '{target_class}' 반에 학생 데이터가 없습니다.")
                        selected_students = []
                else:
                    st.warning("⚠ 배정된 반 이름이 확인되지 않습니다.")
                    target_class = st.text_input("출력할 반 이름 직접 입력:", placeholder="예: S반", key="batch_class_custom")
                    selected_students = None
            else:
                st.warning(f"⚠ [{selected_quarter}]로 등록된 시험 과정이 없습니다.")
                target_class = ""; selected_students = None; selected_test_batch = ""
        else:
            st.warning("⚠ 분석 가능한 분기 명칭이 존재하지 않습니다.")
            target_class = ""; selected_students = None; selected_test_batch = ""
    else:
        st.warning("⚠ 성적 데이터베이스가 비어 있습니다.")
        target_class = ""; selected_students = None; selected_test_batch = ""

    if st.button("반 전체/선택 일괄 생성 (ZIP)", type="primary", key="batch_zip_btn"):
        if not target_class or not target_class.strip():
            st.error("반 이름을 올바르게 선택하거나 기입해 주십시오.")
        elif not selected_test_batch:
            st.error("처리 대상 시험 과정이 지정되지 않았습니다.")
        elif selected_students is not None and len(selected_students) == 0:
            st.error("출력할 학생을 한 명 이상 선택해 주세요.")
        else:
            with st.spinner(f"[{selected_quarter} - {selected_test_batch}] '{target_class}' 반 리포트 생성 중..."):
                success, buf, msg = generate_batch_report(target_class, selected_test_batch, selected_students)
                if success:
                    st.success(msg)
                    st.download_button(
                        "📥 일괄 다운로드 (ZIP)", data=buf.getvalue(),
                        file_name=f"{selected_quarter}_{selected_test_batch}_{target_class}_리포트_패키지.zip",
                        mime="application/zip"
                    )
                else: st.error(msg)

# --- Tab 4: 분기별 엑셀 추출 ---
with tab4:
    st.subheader("📚 분기별 재원생 성적 데이터 엑셀 내보내기 (반별 시트 구성)")
    st.markdown("선택한 **분기**에서 구분이 **'재원생'**인 학생 데이터를 **반명별 개별 시트**로 나누어 엑셀을 자동 마스터링합니다.")
    
    if not df_results_all.empty and '분기' in df_results_all.columns:
        quarter_options = sorted(df_results_all['분기'].dropna().astype(str).unique().tolist())
    else:
        quarter_options = ["2026년 1분기","2026년 2분기","2026년 3분기","2026년 4분기"]
        
    excel_quarter = st.selectbox("📥 내보낼 분기를 선택하세요:", quarter_options, key="excel_quarter_select")
    
    if st.button("📊 해당 분기 재원생 통합 엑셀 파일 생성하기", type="primary"):
        filtered_df = df_results_all[
            (df_results_all['분기'].astype(str).str.strip() == excel_quarter.strip()) &
            (df_results_all['구분'].astype(str).str.strip() == '재원생')
        ].copy()
        
        if filtered_df.empty:
            st.warning(f"⚠ [{excel_quarter}] 분기에 '재원생' 데이터가 없습니다.")
        else:
            distinct_tests     = filtered_df['시험명'].dropna().unique().tolist()
            df_info_filtered_e = df_info_all[df_info_all['시험명'].isin(distinct_tests)]
            
            def clean_info_q(q):
                nums = re.findall(r'\d+', str(q).split('.')[0])
                return nums[0] if nums else str(q).strip()
                
            if not df_info_filtered_e.empty:
                actual_q_cols = df_info_filtered_e['문항번호'].apply(clean_info_q).unique().tolist()
                actual_q_cols = sorted(actual_q_cols, key=lambda x: int(re.findall(r'\d+',x)[0]) if re.findall(r'\d+',x) else x)
            else:
                actual_q_cols = sorted([col for col in filtered_df.columns if col.isdigit() and int(col)<=50], key=int)
                
            available_classes = (filtered_df['반'].astype(str).str.strip()
                                 .replace({'0':'미지정','0.0':'미지정','nan':'미지정','':'미지정'})
                                 .unique().tolist())
            st.success(f"🎯 [{excel_quarter}] 재원생 총 {len(filtered_df)}명 확인! 반: {', '.join([f'[{c}]' for c in sorted(available_classes)])}")
            
            with st.spinner("반별 엑셀 생성 중..."):
                excel_file = export_excel_styled(filtered_df, excel_quarter, actual_q_cols)
            
            st.download_button(
                "📥 반별 탭 분할 엑셀(.xlsx) 다운로드",
                data=excel_file.getvalue(),
                file_name=f"JEET_분할_{excel_quarter}_재원생성적.xlsx".replace(" ","_"),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

# --- Tab 5: 성적 수정/삭제 ---
with tab5:
    st.subheader("✏️ 학생 성적 수정 / 삭제")
    st.markdown("검색 후 해당 학생의 성적을 수정하거나 레코드를 삭제할 수 있습니다.")

    sc1, sc2, sc3 = st.columns([2, 2, 1])
    with sc1:
        edit_test = st.selectbox("시험 과정 선택",
                                 df_info_all['시험명'].dropna().unique().tolist(),
                                 key="edit_test_select")
    with sc2:
        edit_name = st.text_input("학생 이름 입력", placeholder="예: 홍길동", key="edit_name_input")
    with sc3:
        st.markdown("<br>", unsafe_allow_html=True)
        search_btn = st.button("🔍 검색", use_container_width=True, key="edit_search_btn")

    if search_btn:
        if not edit_name.strip():
            st.error("학생 이름을 입력해주세요.")
        else:
            search_name_clean = edit_name.strip().replace(" ","").upper()
            df_edit_pool = df_results_all[
                df_results_all['시험명'].astype(str).str.strip() == str(edit_test).strip()
            ].copy()
            df_edit_pool['이름_정제'] = df_edit_pool['이름'].astype(str).str.replace(" ","").str.upper()
            df_found = df_edit_pool[df_edit_pool['이름_정제'] == search_name_clean]
            
            if df_found.empty:
                st.warning(f"[{edit_test}] 과정에서 '{edit_name}' 학생을 찾을 수 없습니다.")
            else:
                st.session_state["edit_results"] = df_found.to_dict("records")
                st.session_state["edit_test"]    = edit_test

    if "edit_results" in st.session_state and st.session_state.get("edit_test") == edit_test:
        records = st.session_state["edit_results"]
        if len(records) > 1:
            st.info(f"동일 이름의 레코드가 {len(records)}개 있습니다.")
            record_labels = [
                f"[{r.get('분기','?')}] {r.get('이름','')} | {r.get('반','')}반 | 총점: {r.get('총점','')} | id: {r.get('id','')}"
                for r in records
            ]
            selected_label = st.radio("수정할 레코드 선택:", record_labels, key="edit_record_select")
            target_record  = records[record_labels.index(selected_label)]
        else:
            target_record = records[0]

        st.markdown("---")
        st.markdown(f"### 📝 수정 폼 — {target_record.get('이름','')} 학생 (id: `{target_record.get('id','')}`)")

        m1,m2,m3,m4,m5,m6 = st.columns([1.2,1.5,1.5,1.5,1.2,1.5])
        with m1:
            new_type = st.radio("구분", ["재원생","신규생"],
                                index=["재원생","신규생"].index(target_record.get("구분","재원생"))
                                      if target_record.get("구분","재원생") in ["재원생","신규생"] else 0,
                                key="edit_type", horizontal=True)
        with m2: new_name   = st.text_input("이름",  value=target_record.get("이름",""),  key="edit_name_field")
        with m3: new_class  = st.text_input("반",    value=target_record.get("반",""),    key="edit_class_field")
        with m4: new_school = st.text_input("학교",  value=target_record.get("학교",""),  key="edit_school_field")
        with m5:
            grade_options = ["중1","중2","중3"]
            cur_grade = str(target_record.get("학년","중1"))
            new_grade = st.selectbox("학년", grade_options,
                                     index=grade_options.index(cur_grade) if cur_grade in grade_options else 0,
                                     key="edit_grade_field")
        with m6:
            quarter_options_edit = ["2025년 4분기","2026년 1분기","2026년 2분기","2026년 3분기","2026년 4분기",
                                    "2027년 1분기","2027년 2분기","기타/정기 평가"]
            cur_quarter = str(target_record.get("분기","2026년 1분기"))
            new_quarter = st.selectbox("분기", quarter_options_edit,
                                       index=quarter_options_edit.index(cur_quarter) if cur_quarter in quarter_options_edit else 0,
                                       key="edit_quarter_field")

        st.markdown("---")
        st.markdown("#### 📋 문항별 정답 수정")
        df_info_edit = df_info_all[df_info_all['시험명'].astype(str).str.strip() == str(edit_test).strip()].copy()
        def clean_q(q):
            nums = re.findall(r'\d+', str(q).split('.')[0])
            return nums[0] if nums else str(q).strip()
        df_info_edit['문항번호']   = df_info_edit['문항번호'].apply(clean_q)
        q_weight_map_edit    = dict(zip(df_info_edit['문항번호'].astype(str), df_info_edit['배점'].astype(int)))
        question_numbers_edit = sorted(q_weight_map_edit.keys(),
                                       key=lambda x: int(re.findall(r'\d+',x)[0]) if re.findall(r'\d+',x) else x)

        new_answers = {}
        if question_numbers_edit:
            for i in range(0, len(question_numbers_edit), 4):
                cols = st.columns(4)
                for j, q_num in enumerate(question_numbers_edit[i:i+4]):
                    with cols[j]:
                        cur_val = target_record.get(str(q_num), 0)
                        try:    cur_ox = "O" if int(float(str(cur_val))) == 1 else "X"
                        except: cur_ox = "X"
                        choice = st.radio(f"**{q_num}번 ({q_weight_map_edit[q_num]}점)**",
                                          options=["O","X"],
                                          index=0 if cur_ox == "O" else 1,
                                          horizontal=True, key=f"edit_q_{q_num}")
                        new_answers[str(q_num)] = 1 if choice == "O" else 0

        new_total = sum(new_answers[q] * q_weight_map_edit[q] for q in new_answers)
        new_c2 = sum(1 for q in new_answers if q_weight_map_edit.get(q)==2 and new_answers[q]==1)
        new_c3 = sum(1 for q in new_answers if q_weight_map_edit.get(q)==3 and new_answers[q]==1)
        new_c4 = sum(1 for q in new_answers if q_weight_map_edit.get(q)==4 and new_answers[q]==1)

        st.markdown("---")
        rc1,rc2,rc3,rc4 = st.columns(4)
        with rc1: st.metric("💯 수정 후 총점", f"{new_total} 점")
        with rc2: st.metric("🟢 2점 정답",     f"{new_c2} 개")
        with rc3: st.metric("🔵 3점 정답",     f"{new_c3} 개")
        with rc4: st.metric("🔴 4점 정답",     f"{new_c4} 개")

        st.markdown("---")
        btn1, btn2 = st.columns([1, 1])
        with btn1:
            if st.button("💾 수정 내용 저장", type="primary", use_container_width=True, key="edit_save_btn"):
                try:
                    record_id = target_record.get("id")
                    update_data = {
                        "구분": new_type, "이름": new_name.strip(), "반": new_class.strip(),
                        "학교": new_school.strip(), "학년": new_grade, "분기": new_quarter,
                        "총점": new_total, "맞은개수_2점": new_c2, "맞은개수_3점": new_c3, "맞은개수_4점": new_c4,
                    }
                    update_data.update(new_answers)
                    supabase.table("student_results").update(update_data).eq("id", record_id).execute()
                    st.cache_data.clear()
                    st.success(f"✅ {new_name} 학생 성적 수정 완료! (총점: {new_total}점)")
                    del st.session_state["edit_results"]
                    time.sleep(1.5); st.rerun()
                except Exception as e:
                    st.error(f"저장 중 오류 발생: {e}")

        with btn2:
            if st.button("🗑️ 이 레코드 삭제", type="secondary", use_container_width=True, key="edit_delete_btn"):
                st.session_state["confirm_delete"] = True

        if st.session_state.get("confirm_delete"):
            st.warning(f"⚠️ 정말로 **{target_record.get('이름','')}** 학생의 레코드를 삭제하시겠습니까? 복구 불가합니다.")
            d1, d2 = st.columns(2)
            with d1:
                if st.button("✅ 네, 삭제합니다", type="primary", key="confirm_yes"):
                    try:
                        record_id = target_record.get("id")
                        supabase.table("student_results").delete().eq("id", record_id).execute()
                        st.cache_data.clear()
                        st.success("🗑️ 레코드가 삭제되었습니다.")
                        del st.session_state["edit_results"]
                        st.session_state["confirm_delete"] = False
                        time.sleep(1.5); st.rerun()
                    except Exception as e:
                        st.error(f"삭제 중 오류 발생: {e}")
            with d2:
                if st.button("❌ 취소", key="confirm_no"):
                    st.session_state["confirm_delete"] = False; st.rerun()
